"""build_roadmap.py — module 5 of the audit runner (M3).

The Priority Call List workbook (deliverable #3). Lifted from
ONEinsurance/build_roadmap.py — same tier language, colors, and Start Here
layout the first customer got — generalized to any client file and upgraded
so the tiers reflect VERIFIED phone data: line type comes from RPV instead
of "the field isn't blank", and DNC status gates every call action.

Tiers:
  A  reachable email + callable cell           -> call + email, PRIME
  B  reachable email, no strong phone          -> email first
  C  email dead/risky + callable cell          -> call only, do not email
  D  email dead/risky + callable landline      -> call only, do not email
  Q  spam-complainer / do-not-mail             -> never email
  X  no working channel (or DNC with dead email) -> suppress

Any contact whose phone is on a DNC registry shows DO NOT CALL in its
action regardless of tier. LOOKUP_FAILED phones count as no phone.
"""
from __future__ import annotations

import os
from datetime import date

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NEVER_EMAIL = {"abuse", "do_not_mail", "donotmail", "spamtrap"}
CALLABLE = {"SAFE_TO_CALL", "SAFE_BUT_NAME_MISMATCH"}

NAVY, TEAL = "1F3A5F", "2C7A7B"
TIER_COLORS = {"A": "2C7A7B", "B": "9CCFC9", "C": "F4D06F", "D": "E8A87C", "Q": "D9776F", "X": "B9C2CC"}
TIER_GUIDE = [
    ("A", "Reachable email + verified cell", "Call + email — start here"),
    ("B", "Reachable email, no strong phone", "Email first"),
    ("C", "Email dead or risky, verified cell", "Call only — do not email"),
    ("D", "Email dead or risky, landline", "Call only — do not email"),
    ("Q", "Spam-complainer / do-not-mail", "NEVER email"),
    ("X", "No working channel", "Suppress — stop paying for these"),
]


def _score_and_tier(m: pd.DataFrame) -> pd.DataFrame:
    zb = m["zb_status"].str.lower()
    callable_ = m["call_verdict"].isin(CALLABLE)
    cell = callable_ & (m["line_type"] == "cell")
    landline = callable_ & (m["line_type"] != "cell")
    email_ok = (zb == "valid")
    dnc = m["call_verdict"].astype(str).str.startswith("DO_NOT_CALL")

    email_pts = zb.map(lambda s: 55 if s == "valid" else 0 if s in NEVER_EMAIL else 20 if s in ("catch-all", "unknown") else 10)
    phone_pts = cell.map({True: 45, False: 0}) + landline.map({True: 30, False: 0})
    m["Priority Score"] = email_pts + phone_pts

    def tier(i):
        if zb.iat[i] in NEVER_EMAIL:
            return "Q"
        if email_ok.iat[i]:
            return "A" if cell.iat[i] else "B"
        if cell.iat[i]:
            return "C"
        if landline.iat[i]:
            return "D"
        return "X"

    m["Tier"] = [tier(i) for i in range(len(m))]

    def action(i):
        t = m["Tier"].iat[i]
        base = {
            "A": "Call + email — PRIME",
            "B": "Email first",
            "C": "Call only — do not email",
            "D": "Call only — do not email",
            "Q": "NEVER EMAIL (spam risk)",
            "X": "Suppress",
        }[t]
        if dnc.iat[i]:
            lit = m["call_verdict"].iat[i] == "DO_NOT_CALL_LITIGATOR"
            return ("DO NOT CALL (TCPA litigator)" if lit else "DO NOT CALL — on DNC registry") + (
                " — email OK" if email_ok.iat[i] else " — suppress")
        return base

    m["Action"] = [action(i) for i in range(len(m))]
    return m


def run(frame: pd.DataFrame, paths: dict, cfg: dict, mock: bool = False) -> dict:
    m = _score_and_tier(frame.copy())

    tier_order = {"A": 0, "B": 1, "C": 2, "D": 3, "Q": 4, "X": 5}
    m = m.sort_values(["Tier", "Priority Score"], ascending=[True, False],
                      key=lambda c: c.map(tier_order) if c.name == "Tier" else c).reset_index(drop=True)
    m["Rank"] = range(1, len(m) + 1)

    # Generic display columns: whatever the client's file actually has.
    cols = cfg.get("columns") or {}
    ident = [c for c in (cols.get("name"), cols.get("email"), cols.get("phone"), cols.get("source")) if c and c in m.columns]
    verdict_cols = [c for c in ("zb_status", "line_type", "call_verdict", "recommended_channel") if c in m.columns]
    lead = ["Rank", "Priority Score", "Tier", "Action"]

    call = m[m["Tier"].isin(["A", "B", "C", "D"])][lead + ident + ["line_type", "recommended_channel"]]
    esafe = m[(m["zb_status"].str.lower() == "valid")][lead + ident]
    dnx = m[m["Action"].str.startswith(("DO NOT CALL", "NEVER EMAIL")) | (m["Tier"] == "Q")][
        ["Tier", "Action"] + ident + ["call_verdict", "zb_status"]]
    full = m[lead + ident + verdict_cols + [c for c in m.columns if not c.startswith("_")
             and c not in lead + ident + verdict_cols and c != "Rank"]]

    client = cfg.get("display_name", "Client").replace(" ", "_")
    suffix = "_MOCK" if mock else ""
    dest = paths["deliver"] if not mock else paths["work"]
    out_path = os.path.join(dest, f"{client}_Action_Roadmap{suffix}.xlsx")

    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        pd.DataFrame().to_excel(w, sheet_name="Start Here", index=False)
        call.to_excel(w, sheet_name="Priority Call List", index=False)
        esafe.to_excel(w, sheet_name="Email-Safe List", index=False)
        dnx.to_excel(w, sheet_name="Do Not Contact", index=False)
        full.to_excel(w, sheet_name="Full Scored Data", index=False)

    _style(out_path, cfg, len(full), full.columns.get_loc("Tier") + 1)
    counts = m["Tier"].value_counts().sort_index().to_dict()
    return {"roadmap": out_path, "tiers": counts}


def _style(path: str, cfg: dict, nrows: int, tier_col_idx: int) -> None:
    wb = load_workbook(path)
    thin = Side(style="thin", color="D5DCE4")
    border = Border(thin, thin, thin, thin)
    hfill = PatternFill("solid", fgColor=NAVY)
    hfont = Font("Arial", bold=True, color="FFFFFF", size=10)

    def style_sheet(ws, tier_letter=None):
        for c in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=c)
            h.fill, h.font, h.border = hfill, hfont, border
            h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(c)].width = 16
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 26
        if ws.max_row > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        if tier_letter and ws.max_row > 1:
            rng = f"{tier_letter}2:{tier_letter}{ws.max_row}"
            for tv, col in TIER_COLORS.items():
                ws.conditional_formatting.add(rng, CellIsRule(
                    operator="equal", formula=[f'"{tv}"'],
                    fill=PatternFill("solid", fgColor=col),
                    font=Font("Arial", bold=True, color="FFFFFF" if tv in ("A", "Q") else "1F3A5F")))

    for name, tier_letter in (("Priority Call List", "C"), ("Email-Safe List", "C"),
                              ("Do Not Contact", "A"), ("Full Scored Data", get_column_letter(tier_col_idx))):
        style_sheet(wb[name], tier_letter)

    # Start Here — same layout the first customer got.
    sh = wb["Start Here"]
    sh.sheet_view.showGridLines = False
    for col, wd in zip("ABCDE", [3, 26, 48, 30, 10]):
        sh.column_dimensions[col].width = wd

    def put(r, c, v, **k):
        x = sh.cell(row=r, column=c, value=v)
        x.font = Font("Arial", **k)
        return x

    title = cfg.get("display_name", "Client")
    put(2, 2, f"{title} — Reachability Action Roadmap", bold=True, size=16, color=NAVY)
    put(3, 2, f"Prepared by ReachAudit · reachaudit.com · {date.today():%B %d, %Y} · {nrows:,} records",
        size=9, color="5A6675")
    put(5, 2, "How to use this file", bold=True, size=12, color=TEAL)
    for i, s in enumerate([
        "1.  Open the Priority Call List tab. It is sorted best to worst. Work top down.",
        "2.  Tiers A and B are PRIME: reachable by email. A also has a verified cell — multi-touch those.",
        "3.  Tiers C and D are CALL-ONLY. Email is dead but the phone is verified good. Do not email.",
        "4.  The Do Not Contact tab is your protection: DNC-registry numbers and spam-risk emails. Never touch.",
        "5.  For any email campaign, use the Email-Safe List tab. It excludes every risky address.",
    ]):
        put(6 + i, 2, s, size=10)

    r = 12
    put(r, 2, "Tier guide", bold=True, size=12, color=TEAL)
    r += 1
    for j, h in zip(range(2, 6), ["Tier", "What it means", "Action", "Records"]):
        c = sh.cell(row=r, column=j, value=h)
        c.fill, c.font, c.border = hfill, hfont, border
        c.alignment = Alignment(horizontal="center")
    r += 1
    first = r
    tl = get_column_letter(tier_col_idx)
    for tv, mean, act in TIER_GUIDE:
        t = sh.cell(row=r, column=2, value=tv)
        t.fill = PatternFill("solid", fgColor=TIER_COLORS[tv])
        t.font = Font("Arial", bold=True, color="FFFFFF" if tv in ("A", "Q") else "1F3A5F")
        t.alignment = Alignment(horizontal="center")
        t.border = border
        for col, val in ((3, mean), (4, act)):
            x = sh.cell(row=r, column=col, value=val)
            x.font, x.border = Font("Arial", size=10), border
        e = sh.cell(row=r, column=5, value=f"=COUNTIF('Full Scored Data'!{tl}2:{tl}{nrows + 1},\"{tv}\")")
        e.font, e.border = Font("Arial", size=10), border
        e.alignment = Alignment(horizontal="center")
        r += 1
    put(r + 1, 2, "Score = email reachability (max 55) + verified phone strength (cell 45 / landline 30). "
        "Higher = work sooner.", size=9, color="5A6675")
    wb.save(path)
